import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pdfplumber
import re
import openpyxl
from openpyxl import load_workbook
import os
from tkcalendar import DateEntry

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de PDFs")
        self.root.geometry("800x400")
        self.root.configure(bg='#ADD8E6')  # Color azul claro para el fondo

        self.file_paths = []
        self.check_vars = []
        self.progress = None
        self.frame = tk.Frame(root, padx=20, pady=20, bg='#ADD8E6')
        self.frame.pack(padx=10, pady=10)

        title_label = tk.Label(self.frame, text="Procesador de Reportes", font="Helvetica 16 bold", bg='#ADD8E6')
        title_label.grid(row=0, column=0, columnspan=3, pady=20)

        report_label = tk.Label(self.frame, text="Tipo de reporte:", font="Helvetica 12", bg='#ADD8E6')
        report_label.grid(row=1, column=0, sticky=tk.W)

        self.report_type = tk.IntVar()

        urban_button = tk.Radiobutton(self.frame, text="URBANO", variable=self.report_type, value=1, font="Helvetica 10", command=self.update_report_type, bg='#ADD8E6', activebackground='#FF4500', selectcolor='#ADD8E6')
        urban_button.grid(row=2, column=0, sticky=tk.W)

        rural_button = tk.Radiobutton(self.frame, text="RURAL", variable=self.report_type, value=2, font="Helvetica 10", command=self.update_report_type, bg='#ADD8E6', activebackground='#FF4500', selectcolor='#ADD8E6')
        rural_button.grid(row=3, column=0, sticky=tk.W)

        load_button = tk.Button(self.frame, text="Cargar PDFs", command=self.load_pdfs, font="Helvetica 10", height=2, width=15, bg='#1E90FF', fg='white', activebackground='#FF4500')
        load_button.grid(row=2, column=1, rowspan=2, padx=20)

        date_label = tk.Label(self.frame, text="Fecha de visita:", font="Helvetica 10", bg='#ADD8E6')
        date_label.grid(row=2, column=2, sticky=tk.W)

        self.visit_date = DateEntry(self.frame, date_pattern='yyyy-mm-dd')
        self.visit_date.grid(row=3, column=2, sticky=tk.W)

        process_button = tk.Button(self.frame, text="Procesar", command=self.process_files, font="Helvetica 10", height=2, width=15, bg='#1E90FF', fg='white', activebackground='#FF4500')
        process_button.grid(row=4, column=1, pady=20)

        self.file_frame = tk.Frame(root, padx=20, pady=20, bg='#ADD8E6')
        self.file_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.data_type = None


    def update_report_type(self):
        if self.report_type.get() == 1:
            self.data_type = 'urban'
        elif self.report_type.get() == 2:
            self.data_type = 'rural'
        else:
            self.data_type = None

    def load_existing_data(self, file_name):
        self.existing_data = set()
        if os.path.exists(file_name):
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                unique_key = (row[2], row[4])  # No. de solicitud y Cédula
                self.existing_data.add(unique_key)

    def load_pdfs(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        for file_path in file_paths:
            if file_path not in self.file_paths:
                self.file_paths.append(file_path)
                var = tk.BooleanVar(value=True)
                self.check_vars.append(var)
                check = tk.Checkbutton(self.file_frame, text=file_path, var=var, font="Helvetica 10")
                check.pack()

    def process_files(self):
        if not self.data_type:
            messagebox.showerror("Error", "Por favor, escoja el tipo de reporte.")
            return
        if not self.file_paths:
            messagebox.showerror("Error", "Por favor, seleccione al menos un archivo PDF.")
            return
        if not self.visit_date.get_date():
            messagebox.showerror("Error", "Por favor ingrese la fecha de visita.")
            return
        self.progress = ttk.Progressbar(self.root, length=100, mode='determinate')
        self.progress.pack()
        self.root.update()
        if self.data_type == 'urban':
            self.process_urban_files()
        elif self.data_type == 'rural':
            self.process_rural_files()
        self.progress.destroy()
        messagebox.showinfo("Éxito", "Los archivos han sido procesados exitosamente.")

    def process_urban_files(self):
            file_name = "datos_Urbano.xlsx"
            full_path = os.path.join(os.getcwd(), file_name)
            self.load_existing_data(full_path)
            pdf_counter = 1
            if os.path.exists(full_path):
                response = messagebox.askyesnocancel("Archivo existente", "El archivo 'datos_Urbano.xlsx' ya existe. ¿Desea añadir los datos procesados al archivo existente?")
                if response is None:
                    return
                elif response is True:
                    workbook = openpyxl.load_workbook(file_name)
                    sheet = workbook.active
                else:
                    new_file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
                    if not new_file_name:
                        return
                    file_name = new_file_name
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
            sheet["A1"] = "Número de PDF"
            sheet["B1"] = "Fecha Creación"
            sheet["C1"] = "No. de solicitud"
            sheet["D1"] = "Nombre del solicitante"
            sheet["E1"] = "Documento"
            sheet["F1"] = "Comuna"
            sheet["G1"] = "Dirección"
            sheet["H1"] = "Teléfono"
            sheet["I1"] = "Fecha de visita"
            sheet["J1"] = "Creado Por"
            sheet["K1"] = "Número de Encuestador"
            row = sheet.max_row + 1
            visit_date = self.visit_date.get_date()
            for file_path, var in zip(self.file_paths, self.check_vars):
                if not var.get():
                    continue
                with pdfplumber.open(file_path) as pdf:
                    text = ""
                    for page in pdf.pages:
                        text += page.extract_text()
                    address_phone_match = re.search(r"Dirección\s+Teléfono\s+([\s\S]*?)\s+(3\d+)", text)
                    address = address_phone_match.group(1).strip() if address_phone_match else "Dirección no encontrada"
                    phone = address_phone_match.group(2).strip() if address_phone_match else "Teléfono no encontrado"
                    creation_createdby_match = re.search(r"Fecha Creación\s+(.*?)\s+Creado Por\s+(.*)", text)
                    creation_date = creation_createdby_match.group(1).strip() if creation_createdby_match else "Fecha de creación no encontrada"
                    created_by = creation_createdby_match.group(2).strip() if creation_createdby_match else "Creado Por no encontrado"
                    application_num_match = re.search(r"No. solicitud\s+(.*)\s+(.*)", text)
                    application = application_num_match.group(1).strip() if application_num_match else "Numero de solicitud no encontrado"
                    name_match = re.search(r"Sexo\s+([\s\S]*?)\s+([\s\S]*?)\s+([\s\S]*?)\s+([\s\S]*?)", text)
                    first_name = name_match.group(1).strip() if name_match else "Nombre no encontrado"
                    last_name = name_match.group(3).strip() if name_match else "Apellido no encontrado"
                    document_number_match = re.search(r"CÉDULA DE\s+(.*?)\s+(.*)", text)
                    document = document_number_match.group(1).strip() if document_number_match else "Documento no encontrado"
                    unique_key = (application, document)
                    if unique_key in self.existing_data:
                        continue
                    self.existing_data.add(unique_key)
                    sheet[f"A{row}"] = pdf_counter
                    sheet[f"B{row}"] = creation_date
                    sheet[f"C{row}"] = application
                    sheet[f"D{row}"] = first_name + " " + last_name
                    sheet[f"E{row}"] = document
                    sheet[f"G{row}"] = address
                    sheet[f"H{row}"] = phone
                    sheet[f"I{row}"] = visit_date
                    sheet[f"J{row}"] = created_by
                    pdf_counter += 1

            workbook.save(file_name)


    def process_rural_files(self):
        file_name = "datos_Rural.xlsx"
        full_path = os.path.join(os.getcwd(), file_name)
        self.load_existing_data(full_path)
        pdf_counter = 1
        if os.path.exists(full_path):
            response = messagebox.askyesnocancel("Archivo existente", "El archivo 'datos_Rural.xlsx' ya existe. ¿Desea añadir los datos procesados al archivo existente?")
            if response is None:
                return
            elif response is True:
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
                pdf_counter = sheet.max_row
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
        sheet["A1"] = "Número de PDF"
        sheet["B1"] = "Fecha Creación"
        sheet["C1"] = "No. de solicitud"
        sheet["D1"] = "Nombre del solicitante"
        sheet["E1"] = "Documento"
        sheet["F1"] = "Corregimiento"
        sheet["G1"] = "Vereda"
        sheet["H1"] = "Teléfono"
        sheet["I1"] = "Fecha de visita"
        sheet["J1"] = "Creado Por"
        sheet["K1"] = "Número de Encuestador"
        row = sheet.max_row + 1
        visit_date = self.visit_date.get_date()
        for file_path, var in zip(self.file_paths, self.check_vars):
            if not var.get():
                continue
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text()
                address_phone_match = re.search(r"Dirección\s+Teléfono\s+([\s\S]*?)\s+(3\d+)", text)
                address = address_phone_match.group(1).strip() if address_phone_match else "Dirección no encontrada"
                phone = address_phone_match.group(2).strip() if address_phone_match else "Teléfono no encontrado"
                creation_createdby_match = re.search(r"Fecha Creación\s+(.*?)\s+Creado Por\s+(.*)", text)
                creation_date = creation_createdby_match.group(1).strip() if creation_createdby_match else "Fecha de creación no encontrada"
                created_by = creation_createdby_match.group(2).strip() if creation_createdby_match else "Creado Por no encontrado"
                application_num_match = re.search(r"No. solicitud\s+(.*)\s+(.*)", text)
                application = application_num_match.group(1).strip() if application_num_match else "Numero de solicitud no encontrado"
                name_match = re.search(r"Sexo\s+([\s\S]*?)\s+([\s\S]*?)\s+([\s\S]*?)\s+([\s\S]*?)", text)
                first_name = name_match.group(1).strip() if name_match else "Nombre no encontrado"
                last_name = name_match.group(3).strip() if name_match else "Apellido no encontrado"
                document_number_match = re.search(r"CÉDULA DE\s+(.*?)\s+(.*)", text)
                document = document_number_match.group(1).strip() if document_number_match else "Documento no encontrado"
                unique_key = (application, document)
                if unique_key in self.existing_data:
                    continue
                self.existing_data.add(unique_key)
                sheet[f"A{row}"] = pdf_counter
                sheet[f"B{row}"] = creation_date
                sheet[f"C{row}"] = application
                sheet[f"D{row}"] = first_name + " " + last_name
                sheet[f"E{row}"] = document
                sheet[f"G{row}"] = address
                sheet[f"H{row}"] = phone
                sheet[f"I{row}"] = visit_date
                sheet[f"J{row}"] = created_by
                row += 1
                pdf_counter += 1
        workbook.save(file_name)

root = tk.Tk()
app = App(root)
root.mainloop()
